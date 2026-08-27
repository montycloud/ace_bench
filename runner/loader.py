"""
Loads scenario definitions from CloudOps_Scenario_Pack.xlsx.
Each scenario provides the benchmark objective (verbatim from CSV) as input to the agent.
"""

import re
import csv
from pathlib import Path

_SCENARIO_DIR = Path(__file__).parent.parent / "benchmark" / "scenarios"
XLSX_PATH = _SCENARIO_DIR / "CloudOps_Scenario_Pack.xlsx"

# System contexts — versioned so changes are traceable
PROMPT_VERSION = "v4"

# Scenarios that cannot be reproduced from a fresh CloudFormation deploy: they
# depend on real account usage history (Cost Explorer, Rightsizing, Savings Plans /
# RI, Compute Optimizer, sustainability data) that CFN cannot fabricate. Excluded
# from the agent-agnostic `toolloop` suite for v1. See docs/ARCHITECTURE.md.
TOOLLOOP_EXCLUSIONS = {
    "COPS-unactioned-rightsizing-opportu-L1-6179a5be",
    "COPS-ranked-savings-roadmap-from-ri-L2-182929a6",
    "COPS-sustainability-improvement-roa-L2-5a54662a",
    "COPS-cost-anomaly-and-allocation-up-L2-035b015f",
    "COPS-efficiency-review-for-request-L2-30d1c7a0",
    "COPS-self-managed-compute-where-man-L1-cdd04d0c",
}


def is_toolloop_runnable(csv_id: str) -> bool:
    """False for scenarios that need real usage history (excluded from toolloop v1)."""
    return csv_id not in TOOLLOOP_EXCLUSIONS

OUTPUT_CONTRACT_L2 = """
Complete your full investigation first, then format your final answer as valid JSON matching this schema exactly. Do not cut the investigation short to produce output faster.
"observations" and "plan" MUST be arrays — include one object per distinct area or action, never collapse unrelated items.
If you cannot verify something from tool output, say so explicitly in "platform_gaps" rather than fabricating details.

{
  "summary": "one paragraph: what was investigated, what the tools surfaced, and the key gaps or risks found",
  "observations": [
    {
      "area": "service or domain area (e.g. IAM, Lambda, S3, VPC, Cost)",
      "finding": "what was found — grounded in what tools actually returned, not inferred",
      "evidence": "direct citation from tool output (tool name and specific data)",
      "severity": "Critical | High | Medium | Low"
    }
  ],
  "platform_gaps": [
    "specific description of what could not be assessed and which tool limitation prevented it"
  ],
  "plan": [
    {
      "action": "specific recommendation",
      "scope": "which resources or service areas this applies to",
      "priority": "High | Medium | Low",
      "effort": "Small | Medium | Large",
      "rationale": "why this action — tied directly to an observation above"
    }
  ]
}"""

OUTPUT_CONTRACT = """
Complete your full investigation first, then format your final answer as valid JSON matching this schema exactly. Do not cut the investigation short to produce output faster.
"findings" and "improvement_plan" MUST be arrays — include one object per resource, never collapse multiple resources into one entry.

{
  "summary": "one paragraph executive summary of what was found",
  "findings": [
    {
      "resource_id": "exact AWS resource name (not ARN)",
      "resource_type": "S3 | Lambda | SecurityGroup | CloudWatch Alarm | ...",
      "check": "the specific compliance check that failed",
      "severity": "Critical | High | Medium | Low",
      "evidence": "cited data from the assessment supporting this finding"
    }
  ],
  "improvement_plan": [
    {
      "resource_id": "exact AWS resource name matching a finding above",
      "action": "specific remediation step",
      "priority": "High | Medium | Low",
      "effort": "Small | Medium | Large"
    }
  ]
}"""

LEGACY_SYSTEM_CONTEXT_L1 = (
    "Important: Complete every investigation in a single run without pausing to ask "
    "for confirmation between steps. Do not say 'Ready to continue?' or wait for approval. "
    "Execute all steps and return the full result in one response. "
    "This is a read-only assessment — do not make any changes to resources."
)

LEGACY_SYSTEM_CONTEXT_L2 = (
    "Important: Complete every task in a single run without pausing. "
    "Do not say 'Ready to continue?' or ask for confirmation — proceed through all steps automatically. "
    "This is a read-only assessment — do not make any changes to resources. "
    "The assessment contains automated compliance findings that exist independently of WAFR questionnaire completion — "
    "a PENDING questionnaire does NOT mean there are no findings. "
    "Use get_findings_summary and list_assessment_top_entities to surface compliance findings. "
    "Cite specific evidence from tool output for every recommendation. "
    "You MUST output the final JSON block at the end of your response no matter what."
)


def _strip_fixture_refs(text: str) -> str:
    """Remove env:fx/, packet:cp/, and read-only constraints for L2 from objective text."""
    text = re.sub(r'from env:fx/\S+', '', text)
    text = re.sub(r'\|\s*packet:cp/\S+', '', text)
    return text.strip().rstrip(';').strip()


def _strip_readonly_constraint(text: str) -> str:
    """For L2, remove the zero-writes constraint from the objective."""
    text = re.sub(r'and zero unapproved environment writes\.?', '', text)
    text = re.sub(r'zero unapproved environment writes\.?', '', text)
    return text.strip()


def load_scenarios(level: int = 1) -> list[dict]:
    """
    Load all scenarios for a given level from the xlsx.

    Returns list of dicts with:
        csv_id          - stable scenario ID from xlsx
        scenario        - scenario name
        category        - top-level category
        subcategory     - subcategory
        level           - 1 or 2
        objective_raw   - verbatim benchmark objective from xlsx
        objective_clean - objective with env:fx/packet:cp stripped (sent to the agent)
        prompt          - full prompt sent to the agent (system context + clean objective)
        prompt_version  - version string for reproducibility
        gold_path       - path to gold.py if it exists
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb['CloudOps Scenario Master']
        headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]

        scenarios = []
        for r in range(3, ws.max_row + 1):
            row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
            if not row.get('Scenario'):
                continue
            if row.get('Level') != level:
                continue

            csv_id    = row.get('Scenario ID (stable)', '')
            objective = row.get('Benchmark Objective', '')
            clean     = _strip_fixture_refs(objective)
            if level == 2:
                prompt = f"{LEGACY_SYSTEM_CONTEXT_L2}\n\n{clean}\n{OUTPUT_CONTRACT_L2}"
            else:
                prompt = f"{LEGACY_SYSTEM_CONTEXT_L1}\n\n{clean}\n{OUTPUT_CONTRACT}"
            # Note: LEGACY_SYSTEM_CONTEXT_* and OUTPUT_CONTRACT* are kept only to populate the
            # saved-result metadata. The agent-agnostic path uses `task` + `output_contract`.

            # look for a gold file
            safe_id   = csv_id.replace('-', '_')
            gold_path = Path(__file__).parent.parent / 'benchmark' / 'gold_labels' / f'{safe_id}.py'

            from runner.contracts import output_contract as _agnostic_contract

            scenarios.append({
                'csv_id':              csv_id,
                'scenario':            row.get('Scenario', ''),
                'category':            row.get('Category', ''),
                'subcategory':         row.get('Subcategory', ''),
                'complexity':          row.get('Complexity', ''),
                'success_criteria':    row.get('Expected Recommendation & Success Criteria', ''),
                'hard_fail':           row.get('Hard-Fail Conditions', ''),
                'wafr_pillar':         row.get('WAFR Pillar Mapping', ''),
                'wafr_phase':          row.get('WAFR Phase Mapping', ''),
                'level':               level,
                'objective_raw':       objective,
                'objective_clean':     clean,
                # ── agent-agnostic fields (used by AgentAdapter.build_input) ──
                'task':                clean,
                'output_contract':     _agnostic_contract(level),
                'toolloop_runnable':   is_toolloop_runnable(csv_id),
                # ── legacy fields (kept for saved-result metadata / back-compat) ──
                'prompt':              prompt,
                'prompt_version':      PROMPT_VERSION,
                'gold_path':           str(gold_path) if gold_path.exists() else None,
            })

        return scenarios

    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")


def load_scenario(csv_id: str) -> dict:
    """Load a single scenario by its CSV ID."""
    all_scenarios = load_scenarios(level=1) + load_scenarios(level=2)
    match = next((s for s in all_scenarios if s['csv_id'] == csv_id), None)
    if not match:
        raise ValueError(f"Scenario not found: {csv_id}")
    return match


_OVERLAY_CACHE = None


def _load_overlay() -> dict:
    """Load the agent-agnostic overlay (handles + AWS-native tools), cached."""
    global _OVERLAY_CACHE
    if _OVERLAY_CACHE is None:
        overlay_path = Path(__file__).parent.parent / 'benchmark' / 'gold_labels' / 'agnostic_overlay.py'
        if not overlay_path.exists():
            _OVERLAY_CACHE = {}
        else:
            import importlib.util
            spec = importlib.util.spec_from_file_location('agnostic_overlay', overlay_path)
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            _OVERLAY_CACHE = getattr(module, 'OVERLAY', {})
    return _OVERLAY_CACHE


def load_gold(csv_id: str) -> dict:
    """
    Load gold labels for a scenario, merging the agent-agnostic overlay on top.

    The base label keeps its legacy fields (day2 `expected_tool_calls`, original
    physical-ID `correct_resources`) for historical re-eval.
    The overlay adds the agnostic `correct_resources`/`should_not_flag` handles and
    AWS-native `expected_tools` used by the toolloop path. Only keys present in the
    overlay entry override the base.
    """
    scenario = load_scenario(csv_id)
    if not scenario['gold_path']:
        return {}
    import importlib.util
    spec   = importlib.util.spec_from_file_location('gold', scenario['gold_path'])
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    gold = dict(getattr(module, 'GOLD', {}))

    overlay = _load_overlay().get(csv_id)
    if overlay:
        gold.update(overlay)  # overlay wins for correct_resources / should_not_flag / expected_tools
    return gold


if __name__ == '__main__':
    scenarios = load_scenarios(level=1)
    print(f"Loaded {len(scenarios)} L1 scenarios\n")
    for s in scenarios:
        gold = '✓ gold' if s['gold_path'] else '  no gold'
        print(f"  {gold}  {s['csv_id']}")
        print(f"          {s['scenario']}")
        print(f"          clean: {s['objective_clean'][:80]}...")
        print()
